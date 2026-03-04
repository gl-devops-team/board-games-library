"""
Utility functions for processing board game data.

This module provides helper functions for reading and transforming
data from Excel files into JSON-serializable Python structures.
"""

from openpyxl import load_workbook


def get_games_from_xlsx(file_path: str) -> list[dict]:
    """
    Load board game data from an Excel (.xlsx) file.

    The function:
    - Opens the Excel workbook from the provided file path
    - Reads the first (active) worksheet
    - Extracts column headers from the first row
    - Converts subsequent rows into dictionaries
    - Maps selected columns into a standardized game structure

    Args:
        file_path (str):
            Path to the Excel file containing board game data.

    Returns:
        list[dict]:
            A list of dictionaries representing board games.
            Each dictionary contains:

            - name (str | None)
            - players (str | None)
            - time (str | None)
            - description (str | None)
            - image_url (str)

    Notes:
        - Cell values are converted to strings to prevent Excel
          auto-formatting issues (e.g., "2-10" being interpreted as a date).
        - Empty cells are converted to empty strings.
        - The first worksheet (active sheet) is used by default.

    Raises:
        FileNotFoundError:
            If the provided file path does not exist.

        openpyxl.utils.exceptions.InvalidFileException:
            If the file is not a valid Excel file.
    """

    # Load the Excel workbook from the given file path
    workbook = load_workbook(filename=file_path)

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    # Read the header row (first row) to get column names
    headers = [cell.value for cell in sheet[1]]

    # Initialize an empty list to store game dictionaries
    games: list[dict] = []

    # Iterate over rows starting from the second row (skip headers)
    # Use values_only=False to get Cell objects instead of processed values
    for row in sheet.iter_rows(min_row=2, values_only=False):
        row_data: dict = {}

        # Map each header to the corresponding cell value
        for header, cell in zip(headers, row, strict=False):

            # Convert cell value to string to prevent Excel auto-formatting
            # issues (e.g., "2-10" being converted to a date).
            # If the cell is empty (None), use an empty string.
            row_data[header] = str(cell.value) if cell.value is not None else ""

        # Create a dictionary for each game using selected columns
        games.append(
            {
                "name": row_data.get("Name"),
                "players": row_data.get("Number of players"),
                "time": row_data.get("Average game time [h:mm]"),
                "description": row_data.get("Category"),
                "image_url": row_data.get("Image URL", ""),
            }
        )

    # Return the list of game dictionaries
    return games