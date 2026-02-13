from openpyxl import load_workbook


def get_games_from_xlsx(file_path: str):
    # Load the Excel workbook
    workbook = load_workbook(filename=file_path)

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    # Read header row (first row in the sheet)
    headers = [cell.value for cell in sheet[1]]

    games = []

    # Iterate over rows starting from the second row (skip headers)
    # values_only=True returns cell values instead of cell objects
    for row in sheet.iter_rows(min_row=2, values_only=True):

        # Map header names to row values
        row_data = dict(zip(headers, row))

        # Create a dictionary for each game
        games.append({
            'name': row_data.get('Name'),
            'players': row_data.get('Number of players'),
            'time': row_data.get('Average game time [h:mm]'),
            'description': row_data.get('Category')
        })

    return games
